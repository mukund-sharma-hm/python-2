#4. Use the python Faker module to generate fake data for the following.
#	a. Create an excel sheet "Employee Personal Details" with following data. Generate around 50-100 records
#		"EMP ID", "EMP NAME", "EMP EMAIL", "Businees Unit" "Salary"
#
#	4a. WAF to return the empolyee name with top most salary
#	4b. WAF to return the Business Unit with top most aggregated salary
#	4c. WAF to return the employee name in each Business Unit with top most salary
#	4d. WAF Delete the Record of the Employee name from the Excel File with the least salary.
#	4e. WAF to Update the Salary Details of an Employee in the Excel File
    
from faker import Faker
import openpyxl
import random

def fake_employee_data():
    """function to create fake data"""
    fake = Faker()
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(["Employee ID", "Employee Name", "Email", "Business Unit", "Salary"])
    for i in range(random.randint(50,100)):
        ws.append([
            fake.uuid4(),
            fake.name(),
            fake.email(),
            fake.random.choice(["IT","Finance","Marketing","HR"]),
            fake.random_int(min=25000, max=400000)
        ])
    wb.save("Employee_Details.xlsx")
fake_employee_data();

def get_emp_with_max_salary():
    wb = openpyxl.load_workbook('Employee_Details.xlsx')
    ws = wb.active

    max_salary = max(cell for row in ws.iter_rows(min_row=2, min_col=5, max_col=5, values_only=True)for cell in row)
    print(max_salary)
get_emp_with_max_salary()

def business_unit_with_highest_aggregated_salary():

    aggragated_salaries = {}
    wb = openpyxl.load_workbook("Employee_Details.xlsx")
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        business_unit = row[3]
        salary = row[4]

        if business_unit in aggragated_salaries:
            aggragated_salaries[business_unit] += salary
        else:
            aggragated_salaries[business_unit] = salary
    
    max_aggragated_salary = max(aggragated_salaries.values())

    #this is using list comprehension

    # max_aggragated_salary_business_unit = [bu for bu, salary in aggragated_salaries.items() if salary == max_aggragated_salary]
    # print(max_aggragated_salary_business_unit[0])

    # I will be using generator expression here

    max_aggragated_salary_business_unit = next(bu for bu,salary in aggragated_salaries.items() if salary == max_aggragated_salary)    
    print(max_aggragated_salary_business_unit)

business_unit_with_highest_aggregated_salary()


def employee_with_highest_salary_per_bu():
    wb = openpyxl.load_workbook("Employee_Details.xlsx")
    ws = wb.active
    highest_salaries_in_bu = {}

    for row in ws.iter_rows(min_row=2, values_only = True):
        employee_name = row[1]
        salary = row[4]
        bu = row[3]

        if bu in  highest_salaries_in_bu:
            if salary > highest_salaries_in_bu[bu][0]:
                highest_salaries_in_bu[bu] = (salary, employee_name)
        else:
            highest_salaries_in_bu[bu] = (salary, employee_name)

    for i in highest_salaries_in_bu.items():
        print(i[0]+"--->"+i[1][1]+"--->"+ str(i[1][0]))

employee_with_highest_salary_per_bu()

def delete_employee_with_least_salary():
    wb = openpyxl.load_workbook("Employee_Details.xlsx")
    ws = wb.active

    least_salary = float('inf')
    row_index = 0


    for current_row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True),start=2):
        salary = row[4]
        if salary is not None and salary<least_salary:
            least_salary = salary
            row_index = current_row_index

    if salary != 0:
        ws.delete_rows(row_index)
        
delete_employee_with_least_salary()

def update_employee_salary(emp_id, new_salary):
        wb = openpyxl.load_workbook("employee_details.xlsx")
        ws = wb.active
        for row in ws.iter_rows(min_row=2,min_col=1, values_only=True):
            if row[0] == emp_id:
                row[4] = new_salary
                break
        wb.save("employee_details.xlsx")

update_employee_salary(emp_id="8d1935bd-593a-4cfe-9e71-bb9b96f93022", new_salary=75000)
